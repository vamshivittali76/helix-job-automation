from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
TRACKER_PATH = Path(__file__).parent.parent.parent / "output" / "tracker.xlsx"

ALL_JOBS_SHEET = "Applications"
APPLIED_SHEET = "Applied"
STATS_SHEET = "Stats"

COLUMNS = [
    ("Job ID", 14),
    ("Company", 22),
    ("Role Title", 32),
    ("Location", 22),
    ("Type", 10),
    ("Source", 10),
    ("Job URL", 40),
    ("Sponsorship", 16),
    ("Match Score", 12),
    ("LLM Score", 10),
    ("Quality Score", 13),
    ("Date Found", 12),
    ("Date Applied", 12),
    ("Status", 14),
    ("Status Source", 13),
    ("Last Updated", 12),
    ("Resume File", 30),
    ("Cover Letter", 30),
    ("Notes", 30),
]

STATUS_COLORS = {
    "new": "4472C4",
    "applied": "FFC000",
    "interviewing": "70AD47",
    "assessment": "9DC3E6",
    "rejected": "FF6B6B",
    "offer": "00B050",
    "ghosted": "A6A6A6",
    "skipped": "D9D9D9",
}

SPONSORSHIP_COLORS = {
    "sponsor_likely": "C6EFCE",
    "sponsor_unlikely": "FFC7CE",
    "unknown": "FFEB9C",
}

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_row(ws) -> None:
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _job_row_values(job: dict) -> list:
    return [
        job.get("id", ""),
        job.get("company", ""),
        job.get("title", ""),
        job.get("location", ""),
        job.get("remote_type", ""),
        job.get("source", ""),
        job.get("url", ""),
        job.get("sponsorship_status", "unknown"),
        job.get("match_score", 0),
        job.get("llm_score", ""),
        job.get("quality_score", ""),
        _format_date(job.get("date_found")),
        _format_date(job.get("date_applied")),
        job.get("application_status", "new"),
        job.get("status_source", ""),
        _format_date(job.get("last_status_update")),
        job.get("resume_path", ""),
        job.get("cover_letter_path", ""),
        job.get("notes", ""),
    ]


def _write_job_rows(ws, jobs: list[dict], start_row: int = 2) -> None:
    for offset, job in enumerate(jobs):
        row = start_row + offset
        values = _job_row_values(job)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        _apply_row_formatting(ws, row)


def _create_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = ALL_JOBS_SHEET
    _apply_header_row(ws)

    applied_ws = wb.create_sheet(APPLIED_SHEET, index=1)
    _apply_header_row(applied_ws)

    stats_ws = wb.create_sheet(STATS_SHEET)
    stats_ws["A1"] = "Metric"
    stats_ws["B1"] = "Value"
    stats_ws["A1"].font = Font(bold=True)
    stats_ws["B1"].font = Font(bold=True)

    return wb


def _get_workbook() -> Workbook:
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    if TRACKER_PATH.exists():
        return load_workbook(str(TRACKER_PATH))
    wb = _create_workbook()
    wb.save(str(TRACKER_PATH))
    return wb


def _rebuild_applied_sheet(wb: Workbook, jobs: list[dict]) -> None:
    """Replace the Applied sheet with a full snapshot of submitted applications."""
    if APPLIED_SHEET in wb.sheetnames:
        wb.remove(wb[APPLIED_SHEET])
    insert_idx = wb.sheetnames.index(STATS_SHEET) if STATS_SHEET in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet(APPLIED_SHEET, insert_idx)
    _apply_header_row(ws)
    _write_job_rows(ws, jobs, start_row=2)


def sync_applied_sheet_workbook(wb: Workbook, jobs: list[dict]) -> None:
    """Refresh only the Applied sheet from a pre-fetched job list."""
    _rebuild_applied_sheet(wb, jobs)


def sync_applied_sheet() -> None:
    """Refresh Applied sheet from DB and save tracker.xlsx (used after each successful apply)."""
    from src.tracker.db import get_applied_jobs

    wb = _get_workbook()
    sync_applied_sheet_workbook(wb, get_applied_jobs(limit=10000))
    wb.save(str(TRACKER_PATH))


def export_jobs_to_excel(jobs: list[dict]):
    """Write/update all jobs to the Applications sheet; rebuild Applied sheet."""
    from src.tracker.db import get_applied_jobs

    wb = _get_workbook()
    ws = wb[ALL_JOBS_SHEET]

    existing_ids = set()
    for row in range(2, ws.max_row + 1):
        job_id = ws.cell(row=row, column=1).value
        if job_id:
            existing_ids.add(job_id)

    for job in jobs:
        job_id = job.get("id", "")
        if job_id in existing_ids:
            _update_existing_row(ws, job)
            continue

        row = ws.max_row + 1
        values = _job_row_values(job)

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        _apply_row_formatting(ws, row)

    _update_stats(wb, jobs)
    sync_applied_sheet_workbook(wb, get_applied_jobs(limit=10000))
    wb.save(str(TRACKER_PATH))


def _update_existing_row(ws, job: dict):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == job.get("id"):
            ws.cell(row=row, column=9).value = job.get("match_score", 0)
            if job.get("llm_score"):
                ws.cell(row=row, column=10).value = job["llm_score"]
            if job.get("application_status"):
                ws.cell(row=row, column=14).value = job["application_status"]
            if job.get("date_applied"):
                ws.cell(row=row, column=13).value = _format_date(job["date_applied"])
            if job.get("quality_score"):
                ws.cell(row=row, column=11).value = job["quality_score"]
            _apply_row_formatting(ws, row)
            break


def _apply_row_formatting(ws, row: int):
    # Status column is 14
    status = str(ws.cell(row=row, column=14).value or "new").lower()
    status_color = STATUS_COLORS.get(status, "FFFFFF")
    ws.cell(row=row, column=14).fill = PatternFill(
        start_color=status_color, end_color=status_color, fill_type="solid"
    )
    if status in ("rejected", "offer", "ghosted"):
        ws.cell(row=row, column=14).font = Font(color="FFFFFF", bold=True)

    sponsorship = str(ws.cell(row=row, column=8).value or "unknown").lower()
    sp_color = SPONSORSHIP_COLORS.get(sponsorship, "FFFFFF")
    ws.cell(row=row, column=8).fill = PatternFill(
        start_color=sp_color, end_color=sp_color, fill_type="solid"
    )

    for score_col in (9, 10):
        score = ws.cell(row=row, column=score_col).value
        if score and isinstance(score, (int, float)):
            if score >= 70:
                ws.cell(row=row, column=score_col).font = Font(color="006100", bold=True)
            elif score >= 50:
                ws.cell(row=row, column=score_col).font = Font(color="9C6500")
            else:
                ws.cell(row=row, column=score_col).font = Font(color="9C0006")


def _update_stats(wb: Workbook, jobs: list[dict]):
    if STATS_SHEET not in wb.sheetnames:
        return
    ws = wb[STATS_SHEET]
    for row in range(2, 20):
        ws.cell(row=row, column=1).value = None
        ws.cell(row=row, column=2).value = None

    total = len(jobs)
    applied = sum(1 for j in jobs if j.get("is_applied"))
    sponsor_likely = sum(1 for j in jobs if j.get("sponsorship_status") == "sponsor_likely")
    avg_score = (
        sum(j.get("match_score", 0) for j in jobs) / total if total else 0
    )

    stats = [
        ("Total Jobs Found", total),
        ("Applications Submitted", applied),
        ("Sponsor Likely", sponsor_likely),
        ("Average Match Score", f"{avg_score:.1f}"),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for idx, (metric, value) in enumerate(stats, 2):
        ws.cell(row=idx, column=1).value = metric
        ws.cell(row=idx, column=2).value = value


def _format_date(date_str) -> str:
    if not date_str:
        return ""
    try:
        if "T" in str(date_str):
            dt = datetime.fromisoformat(str(date_str).replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d")
        return str(date_str)[:10]
    except (ValueError, TypeError):
        return str(date_str) if date_str else ""
